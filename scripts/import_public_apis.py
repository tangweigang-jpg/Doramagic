#!/usr/bin/env python3
"""import_public_apis.py — 导入 public-apis 外部参考 catalog 到 YAML

从 public-apis-catalog.xlsx 读取 API 目录，转换为结构化 YAML 写入
knowledge/catalogs/public_apis.yaml。

用法:
    python scripts/import_public_apis.py \\
        --input /path/to/public-apis-catalog.xlsx \\
        --output knowledge/catalogs/public_apis.yaml

输出:
    - knowledge/catalogs/public_apis.yaml（幂等，重跑覆写）
    - stdout 统计摘要

退出码:
    0  成功
    1  输入错误
    2  依赖缺失
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

try:
    import yaml
except ImportError:
    print("[error] PyYAML not installed. Run: pip install pyyaml", file=sys.stderr)
    sys.exit(2)

try:
    import openpyxl
except ImportError:
    print("[error] openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# Categories considered finance-relevant
FINANCE_RELEVANT_CATEGORIES = frozenset(
    [
        "Finance",
        "Cryptocurrency",
        "Business",
        "Currency Exchange",
        "Blockchain",
    ]
)


def load_xlsx(input_path: Path) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Load the xlsx file and return (api_rows, category_summary).

    api_rows — list of dicts from 'All APIs' sheet
    category_summary — dict from '分类汇总' sheet
    """
    wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)

    # ---------- All APIs sheet ----------
    if "All APIs" not in wb.sheetnames:
        print(f"[error] Sheet 'All APIs' not found in {input_path}", file=sys.stderr)
        sys.exit(1)

    ws_apis = wb["All APIs"]
    rows = list(ws_apis.iter_rows(values_only=True))
    if not rows:
        print("[error] 'All APIs' sheet is empty", file=sys.stderr)
        sys.exit(1)

    # Row 1 = header
    headers: list[str] = [str(h) if h is not None else "" for h in rows[0]]

    def _col(name: str) -> int:
        """Return 0-based index of a column by exact or partial match."""
        for i, h in enumerate(headers):
            if name.lower() in h.lower():
                return i
        return -1

    col_category = _col("分类")
    col_name = _col("API 名称")
    col_url = _col("地址")
    col_desc = _col("主要功能")
    col_free = _col("是否免费")
    col_auth = _col("认证方式")
    col_https = _col("HTTPS")
    col_cors = _col("CORS")

    def _get(row: tuple, idx: int) -> Any:
        if idx < 0 or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        if isinstance(v, str) and v.strip() == "":
            return None
        return v

    api_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        category = _get(row, col_category)
        name = _get(row, col_name)
        if not name:
            continue

        # Normalize HTTPS/CORS to bool or None
        https_raw = _get(row, col_https)
        if https_raw is True or (
            isinstance(https_raw, str) and https_raw.strip().lower() == "true"
        ):
            https_val: bool | None = True
        elif https_raw is False or (
            isinstance(https_raw, str) and https_raw.strip().lower() == "false"
        ):
            https_val = False
        else:
            https_val = None

        cors_raw = _get(row, col_cors)
        cors_val = str(cors_raw).strip() if cors_raw is not None else None

        api_rows.append(
            {
                "category": str(category) if category else None,
                "name": str(name).strip(),
                "url": str(_get(row, col_url) or "").strip() or None,
                "description_zh": str(_get(row, col_desc) or "").strip() or None,
                "free_status": str(_get(row, col_free) or "").strip() or None,
                "auth": str(_get(row, col_auth) or "").strip() or None,
                "https": https_val,
                "cors": cors_val,
            }
        )

    # ---------- 分类汇总 sheet ----------
    category_summary: dict[str, int] = {}
    if "分类汇总" in wb.sheetnames:
        ws_cat = wb["分类汇总"]
        for row in ws_cat.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and row[0] and row[1] is not None:
                category_summary[str(row[0]).strip()] = int(row[1])

    wb.close()
    return api_rows, category_summary


def build_catalog(
    api_rows: list[dict[str, Any]],
    category_summary: dict[str, int],
    input_path: Path,
) -> dict[str, Any]:
    """Build the catalog YAML structure."""
    imported_at = datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")

    finance_relevant_count = sum(
        1 for r in api_rows if r.get("category") in FINANCE_RELEVANT_CATEGORIES
    )

    # Sort by category then name
    sorted_apis = sorted(api_rows, key=lambda r: (r.get("category") or "", r.get("name") or ""))

    meta: dict[str, Any] = {
        "source": "public-apis (GitHub community catalog)",
        "imported_from": input_path.name,
        "imported_at": imported_at,
        "total_entries": len(api_rows),
        "categories_count": len(category_summary)
        if category_summary
        else len(set(r.get("category") for r in api_rows)),
        "finance_relevant_count": finance_relevant_count,
        "finance_relevant_categories": sorted(FINANCE_RELEVANT_CATEGORIES),
    }

    return {
        "meta": meta,
        "category_summary": dict(sorted(category_summary.items())),
        "apis": sorted_apis,
    }


def print_stats(catalog: dict[str, Any]) -> None:
    meta = catalog["meta"]
    apis = catalog["apis"]

    print("\n" + "=" * 60)
    print("  Public APIs Catalog Import — Summary")
    print("=" * 60)
    print(f"  Source             : {meta['source']}")
    print(f"  Total entries      : {meta['total_entries']}")
    print(f"  Categories         : {meta['categories_count']}")
    print(f"  Finance-relevant   : {meta['finance_relevant_count']}")
    print(f"    (categories: {', '.join(meta['finance_relevant_categories'])})")

    cat_counts = Counter(r.get("category") for r in apis)
    top10 = cat_counts.most_common(10)
    print("\n  Top 10 categories:")
    for cat, count in top10:
        print(f"    {cat!s:<35} {count}")

    # Finance-relevant breakdown
    print("\n  Finance-relevant category breakdown:")
    for cat in sorted(FINANCE_RELEVANT_CATEGORIES):
        count = cat_counts.get(cat, 0)
        print(f"    {cat:<35} {count}")

    print("=" * 60 + "\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import public-apis-catalog.xlsx to YAML.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input", required=True, metavar="XLSX", help="Path to public-apis-catalog.xlsx"
    )
    parser.add_argument("--output", required=True, metavar="PATH", help="Output YAML path")
    parser.add_argument("--base-dir", default=None, metavar="DIR", help="Project root directory")
    return parser.parse_args()


def resolve_base_dir(args: argparse.Namespace) -> Path:
    if args.base_dir:
        return Path(args.base_dir).resolve()
    script_dir = Path(__file__).resolve().parent
    candidate = script_dir.parent
    if (candidate / "knowledge").exists():
        return candidate
    return Path.cwd()


def main() -> None:
    args = parse_args()
    base_dir = resolve_base_dir(args)

    input_path = Path(args.input)
    if not input_path.is_absolute():
        input_path = base_dir / input_path
    if not input_path.exists():
        print(f"[error] Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = base_dir / output_path

    print(f"[import] Loading {input_path} ...")
    api_rows, category_summary = load_xlsx(input_path)
    print(f"[import] Loaded {len(api_rows)} API rows, {len(category_summary)} categories.")

    catalog = build_catalog(api_rows, category_summary, input_path)
    print_stats(catalog)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        yaml.dump(
            catalog,
            f,
            allow_unicode=True,
            default_flow_style=False,
            sort_keys=False,
            width=120,
        )

    size = output_path.stat().st_size
    print(f"[output] Written to : {output_path}")
    print(f"[output] File size  : {size:,} bytes")


if __name__ == "__main__":
    main()
